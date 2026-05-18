"""
excel_export.py
---------------
Colour-coded Excel output matching the production reporting format.

Zone colours:
  green  (#92D050) — Keine Reklamation  (confident machine problem)
  yellow (#FFFF00) — Unsicher           (model uncertain, review needed)
  red    (#FF0000) — Reklamation        (likely paper problem)
"""

import logging
import pandas as pd

log = logging.getLogger(__name__)

_FILL_COLOURS = {
    "grün": "FF92D050",
    "gelb": "FFFFFF00",
    "rot":  "FFFF0000",
}


def write_coloured_excel(df: pd.DataFrame, path: str) -> None:
    """
    Writes a colour-coded Excel file.
    Silently skips if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    headers     = list(df.columns)
    thin        = Side(style="thin", color="FF000000")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FF4472C4")
    header_font = Font(bold=True, color="FFFFFFFF")
    center      = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=h)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.border     = border
        cell.alignment  = center

    farbe_col   = headers.index("farbe") + 1
    einsch_col  = headers.index("einschaetzung") + 1
    score_col   = headers.index("score_%") + 1
    coloured    = {farbe_col, einsch_col, score_col}

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        farbe = row.farbe
        fill  = PatternFill("solid", fgColor=_FILL_COLOURS.get(farbe, "FFFFFFFF"))
        font_red = Font(bold=True, color="FFFFFFFF") if farbe == "rot" else None

        for col_idx, val in enumerate(row, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = border
            cell.alignment = center
            if col_idx in coloured:
                cell.fill = fill
                if font_red:
                    cell.font = font_red

    _col_widths = {
        "event_id": 28, "source": 22, "fold": 8,
        "true_label": 12, "true_label_str": 18,
        "probability": 12, "prediction": 12, "prediction_str": 18,
        "score_%": 10, "confidence_%": 14, "confident_class": 20,
        "einschaetzung": 24, "farbe": 10, "correct": 10, "threshold": 12,
    }
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _col_widths.get(h, 15)

    ws.freeze_panes = "A2"
    wb.save(path)
    log.info("Excel saved → %s", path)
